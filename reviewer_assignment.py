"""
Reviewer Assignment Module
===========================
Maps taxonomy leaf categories to professional domains and generates one
annotated Excel file per domain for the human expert who reviews it.

Stage A (hierarchical_taxonomy.py) decides *how technologies group*.
This module decides *who reads each group* — and, more importantly, where it is
not confident, so a human sees the doubt instead of inheriting it silently.

─────────────────────────────────────────────────────────────────────────────
REDACTED FOR PUBLICATION
─────────────────────────────────────────────────────────────────────────────
The reviewer taxonomy this module was built against belongs to the client and is
not published. Two things are therefore blanked out below, and only those two:

  • SYSTEM_PROMPT — held the full reviewer taxonomy (domain -> sub-domain ->
    sub-sub-domain, several dozen entries), plus the assignment rules, the
    do's and don'ts, and the strict output contract.
  • The domain names themselves — replaced everywhere by DOMAIN_1 … DOMAIN_4.

Everything else is the code as it actually ran: the two-phase architecture, the
anomaly model, the defensive parsing, the cross-domain referral logic and the
workbook layout. The prompts were the easy part; the structure is the work.
─────────────────────────────────────────────────────────────────────────────

Two design decisions carry this module:

1. Classification and anomaly detection are SEPARATE model calls.
   Asking one call to both assign a batch of leaves and spot the odd ones out
   makes it miss the odd ones out — the second task quietly loses to the first.
   A focused, single-leaf call finds substantially more.

2. An anomaly is defined relative to the ROUTING DECISION, not to semantics.
   Two technologies with different mechanisms that land on the same reviewer are
   not a problem — that reviewer reads both either way. Only a technology
   belonging to a *different* reviewer is a real routing error. Measuring
   semantic oddity instead would flood the reviewer with flags that cost
   attention and change nothing.
"""

import io
import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# REVIEWER DOMAINS  (names redacted — see header)
# ─────────────────────────────────────────────────────────────────────────────

DOMAIN_1 = "DOMAIN_1"   # was: a broad engineering domain owned by one reviewer
DOMAIN_2 = "DOMAIN_2"
DOMAIN_3 = "DOMAIN_3"
DOMAIN_4 = "DOMAIN_4"


# ─────────────────────────────────────────────────────────────────────────────
# AGENT PROMPTS
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """
[REDACTED]

This prompt contained the client's full reviewer taxonomy and the rules for
assigning to it. Its structure, which is the part worth reading:

  • Role framing — expert technology classifier assigning to a reviewer taxonomy.
  • Assignment rules:
      1. Default to EXACTLY ONE sub-domain — the most specific level available.
      2. Exception: a technology that genuinely bridges two different domains
         (truly dual-core, not merely related) also gets a secondary. Used
         sparingly, and the prompt says so explicitly — otherwise the model
         hands out secondaries generously and the cross-domain signal becomes
         worthless.
      3. Always report the parent domain, for grouping.
      4. Decide by CORE MECHANISM, not by application. The worked example in the
         prompt: DNA-based energy storage is biology (core = DNA), not energy.
      5. Read the full leaf description, not just the leaf name.
  • The taxonomy itself — domain -> sub-domain -> sub-sub-domain, several dozen
    leaves, each with a short scope note disambiguating it from its neighbours.
  • Output contract — JSON only, no prose, no questions, response must start with
    '{' and end with '}'. The domain field is constrained to an explicit
    enumeration of the domain names, so the model cannot invent a fifth domain
    or spell an existing one differently.
"""


ANOMALY_SYSTEM_PROMPT = f"""You are reviewing a group of emerging technologies that were clustered together algorithmically. This clustering feeds a ROUTING decision: each group is sent to ONE professional reviewer who owns a broad domain. There are exactly four reviewer domains:

1. {DOMAIN_1} — [REDACTED: one-line scope of this reviewer's domain]
2. {DOMAIN_2} — [REDACTED]
3. {DOMAIN_3} — [REDACTED]
4. {DOMAIN_4} — [REDACTED]

THE KEY PRINCIPLE: what matters for a flag is the REVIEWER, not the narrow sub-topic. Two technologies that differ in mechanism but belong to the SAME reviewer domain are NOT a routing problem — the right reviewer still gets both. Only when a technology belongs to a DIFFERENT reviewer domain is the routing actually wrong. Do not flag a technology merely for being a different sub-topic within the same domain.

Judge by each technology's CORE mechanism — not its application label, and not the group name.

Determine TWO things.

(A) GROUP STATUS — exactly one of:
- "tight": (nearly) all technologies belong to ONE reviewer domain AND share a coherent core mechanism. A clean, well-formed group.
- "diverse": all, or a clear majority, belong to the SAME ONE reviewer domain, but they span clearly different sub-topics with no single tight core. The right reviewer is unambiguous; the grouping is just loose.
- "mixed": the technologies span MULTIPLE reviewer domains with no clear majority — a grab-bag that no single reviewer can own.

(B) ITEM ANOMALIES — list the technologies that stand out. Give each a level:
- "yellow": its core differs from the group's shared theme, BUT it still belongs to the SAME reviewer domain as the majority. (Different sub-topic, same reviewer — minor.)
- "red": it belongs to a DIFFERENT reviewer domain than the majority — routing this one to the group's reviewer is likely wrong.
If GROUP STATUS is "mixed", leave anomalies EMPTY — the whole group already goes to manual routing, so per-item flags add nothing.

Respond with ONLY valid JSON, nothing else:
{{
  "majority_reviewer": "<which ONE of the four domains the group mostly belongs to, or 'none' if mixed>",
  "majority_theme": "<one sentence: the shared core; or 'no single core, one reviewer' for diverse; or 'spans multiple domains' for mixed>",
  "group_status": "tight" | "diverse" | "mixed",
  "anomalies": [{{"name": "exact technology name", "level": "yellow" | "red"}}]
}}

Decision guide:
- One reviewer, one shared core, all belong            -> tight,   anomalies: []
- One reviewer, varied sub-topics                      -> diverse, anomalies: [odd sub-topics as yellow]
- Clear majority reviewer + a stray item from ANOTHER domain -> tight or diverse, anomalies: [that stray item as red]
- No majority reviewer, spans domains                  -> mixed,   anomalies: []
"""


def build_user_prompt(leaves_data: list) -> str:
    lines = ["Classify each emerging technology leaf category to the most specific sub-domain.\n"]
    for i, leaf in enumerate(leaves_data, 1):
        lines.append(f"{i}. Leaf: {leaf['leaf_category']}")
        lines.append(f"   Hierarchy: {leaf['full_path']}")
        if leaf.get("summary"):
            lines.append(f"   Description: {leaf['summary'][:400]}")
        if leaf.get("tech_names"):
            names_str = ", ".join(leaf["tech_names"])
            lines.append(f"   Technologies in this group: {names_str}")
        lines.append("")
    return "\n".join(lines)


def _detect_anomalies_for_leaf(leaf: dict, client, model_id: str) -> dict:
    """Separate focused call: classify the group and flag misrouted items.

    Returns {"group_status": "tight"|"diverse"|"mixed",
             "anomalies": [{"name": str, "level": "yellow"|"red"}],
             "theme": str}.

    The judgment is at the REVIEWER level, not the narrow leaf theme: an item is
    only "red" (routing problem) if it belongs to a DIFFERENT reviewer domain;
    a different sub-topic within the same reviewer's domain is at most "yellow".
    A "mixed" group (spans several reviewer domains) is the worst silent-failure
    case — it goes to manual routing as a whole, so we do not pick per-item flags.
    """
    empty = {"group_status": "tight", "anomalies": [], "theme": ""}
    if not leaf.get("tech_names"):
        return empty
    tech_list = "\n".join(f"- {t}" for t in leaf["tech_names"])
    user_msg = f"""Here is a group of technologies that were clustered together algorithmically.

Technologies in this group:
{tech_list}

Group label (may be imprecise): {leaf['leaf_category']}
Context: {leaf.get('summary', '')[:200]}

(A) Which ONE reviewer domain does the majority belong to, and is the group tight / diverse / mixed?
(B) Which technologies (if any) stand out — and for each, is it the SAME reviewer (yellow) or a DIFFERENT reviewer (red)?"""

    try:
        body = json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 600,
            # temperature 0 → reproducible run-to-run (critical for a deliverable)
            # and far less malformed JSON than the default sampling temperature.
            "temperature": 0,
            "system": ANOMALY_SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": user_msg}],
        })
        response = client.invoke_model(modelId=model_id, body=body)
        raw = json.loads(response["body"].read())["content"][0]["text"].strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        if not raw:
            return empty
        # Parse defensively, in three escalating steps, so a leaf is NEVER
        # silently dropped because the model formatted its JSON imperfectly:
        #   1. raw_decode — reads the first object, ignores trailing prose.
        #   2. first {...} block — handles leading prose.
        #   3. field-level regex salvage — recovers group_status + anomaly
        #      name/level pairs even from malformed JSON (missing comma, stray
        #      quote inside a long technology name → "Expecting ',' delimiter").
        data = None
        try:
            data, _end = json.JSONDecoder().raw_decode(raw)
        except json.JSONDecodeError:
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            if m:
                try:
                    data = json.loads(m.group(0))
                except json.JSONDecodeError:
                    data = None
        if data is None:
            sm = re.search(r'"group_status"\s*:\s*"(\w+)"', raw)
            if not sm:
                print(f"    anomaly check unparseable for {leaf['leaf_category']} — נשמר כ-tight, כדאי בדיקה ידנית")
                return empty
            pairs = re.findall(r'"name"\s*:\s*"([^"]+)"\s*,\s*"level"\s*:\s*"(\w+)"', raw)
            data = {
                "group_status": sm.group(1),
                "majority_theme": "",
                "anomalies": [{"name": n, "level": l} for n, l in pairs],
            }
            print(f"    ⚠️ JSON פגום ב-{leaf['leaf_category']} — שוחזר ({data['group_status']}, {len(pairs)} פריטים)")
        theme = data.get("majority_theme", "")
        status = str(data.get("group_status", "tight")).strip().lower()
        if status not in ("tight", "diverse", "mixed"):
            status = "tight"

        # Normalize anomalies into [{"name","level"}]; tolerate plain-string lists.
        raw_anoms = data.get("anomalies", []) or []
        anomalies = []
        for a in raw_anoms:
            if isinstance(a, dict):
                name = str(a.get("name", "")).strip()
                level = str(a.get("level", "yellow")).strip().lower()
                level = "red" if level == "red" else "yellow"
            else:
                name, level = str(a).strip(), "yellow"
            if name:
                anomalies.append({"name": name, "level": level})

        # Precedence: a mixed group goes to manual routing as a whole — per-item
        # flags add nothing, so drop them.
        if status == "mixed":
            anomalies = []

        if status == "mixed":
            print(f"    ⛔ קבוצה מעורבת (תחומים שונים): {leaf['leaf_category']} — לשיוך ידני")
        elif status == "diverse":
            n_red = sum(1 for a in anomalies if a["level"] == "red")
            n_yel = len(anomalies) - n_red
            parts = []
            if n_yel:
                parts.append(f"{n_yel} חריגות (אותו בודק)")
            if n_red:
                parts.append(f"⛔ {n_red} לא מתאימות (בודק אחר!)")
            extra = (" · " + " · ".join(parts)) if parts else ""
            print(f"    ⚠️ קבוצה מגוונת (אותו בודק): {leaf['leaf_category']}{extra}")
        elif anomalies:
            reds = [a["name"] for a in anomalies if a["level"] == "red"]
            yels = [a["name"] for a in anomalies if a["level"] == "yellow"]
            if reds:
                print(f"    ⛔ לא מתאימות (בודק אחר) ב-{leaf['leaf_category']}: {reds}")
            if yels:
                print(f"    ⚠️ חריגות (אותו בודק) ב-{leaf['leaf_category']}: {yels}")
        return {"group_status": status, "anomalies": anomalies, "theme": theme}
    except Exception as e:
        err = str(e)
        if "line 1 column 1" in err:
            # Empty response — treat as "model found nothing", not a problem.
            return empty
        print(f"    anomaly check failed for {leaf['leaf_category']}: {err}")
        return empty


def _flag_view(info: dict):
    """Normalize a leaf's flag data into (group_status, [{"name","level"}, ...]).

    Tolerant of the legacy shape ({"incoherent": bool, "anomalies": [str]}) so
    old saved/mock domain_maps still render without crashing.
    """
    status = info.get("group_status")
    raw_anoms = info.get("anomalies", []) or []
    if status is None:
        status = "mixed" if info.get("incoherent") else "tight"
    status = str(status).strip().lower()
    if status not in ("tight", "diverse", "mixed"):
        status = "tight"
    anomalies = []
    for a in raw_anoms:
        if isinstance(a, dict):
            name = str(a.get("name", "")).strip()
            level = "red" if str(a.get("level", "yellow")).strip().lower() == "red" else "yellow"
        else:
            name, level = str(a).strip(), "yellow"
        if name:
            anomalies.append({"name": name, "level": level})
    if status == "mixed":
        anomalies = []
    return status, anomalies


def _get_active_text_model(client=None) -> str:
    """Get the active text model ID — reuse what hierarchical_taxonomy already picked."""
    try:
        import hierarchical_taxonomy as _ht
        model = _ht.TEXT_MODEL_ID
        print(f"  Reviewer using model: {model}")
        return model
    except Exception:
        return "us.anthropic.claude-3-7-sonnet-20250219-v1:0"


def assign_leaves_to_domains(
    result_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    client,
    model_id: str = "us.anthropic.claude-3-5-haiku-20241022-v1:0",
) -> dict:
    """
    Call Claude via Bedrock client (same client used in classification).
    Returns: {leaf_category: {primary_domain, primary_sub_domain, ...}}
    """
    leaves = (
        result_df[["leaf_category", "full_path"]]
        .drop_duplicates("leaf_category")
        .to_dict("records")
    )

    summary_map = {}
    if summary_df is not None and "category" in summary_df.columns:
        summary_map = dict(zip(summary_df["category"], summary_df.get("summary", "")))

    # Add tech names and summary to each leaf
    tech_names_map = (
        result_df.groupby("leaf_category")["technology_name"]
        .apply(list)
        .to_dict()
    )
    for leaf in leaves:
        leaf["summary"] = summary_map.get(leaf["leaf_category"], "")
        leaf["tech_names"] = tech_names_map.get(leaf["leaf_category"], [])

    def _call_model(batch):
        user_prompt = build_user_prompt(batch)
        body = json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 4000,
            "temperature": 0,  # reproducible domain assignments run-to-run
            "system": SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": user_prompt}],
        })
        response = client.invoke_model(modelId=model_id, body=body)
        raw = json.loads(response["body"].read())["content"][0]["text"].strip()
        print(f"  Batch response: {len(raw)} chars")
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        if not raw:
            raise ValueError(f"Model returned empty response for batch of {len(batch)} leaves")
        return json.loads(raw)["assignments"]

    # Split into batches of 7 — larger batches measurably lose accuracy.
    BATCH_SIZE = 7
    result = {}
    for i in range(0, len(leaves), BATCH_SIZE):
        batch = leaves[i:i + BATCH_SIZE]
        print(f"  Processing batch {i//BATCH_SIZE + 1}/{-(-len(leaves)//BATCH_SIZE)} ({len(batch)} leaves)...")
        assignments = _call_model(batch)
        for item in assignments:
            result[item["leaf_category"]] = {
                "primary_domain":       item["primary"]["domain"],
                "primary_sub_domain":   item["primary"]["sub_domain"],
                "secondary_domain":     item["secondary"]["domain"]     if item.get("secondary") else None,
                "secondary_sub_domain": item["secondary"]["sub_domain"] if item.get("secondary") else None,
                "reasoning": item.get("reasoning", ""),
                "group_status": "tight",
                "anomalies": [],
            }
    # Second pass: focused per-leaf review — group status + misrouted items
    print(f"  Phase 2: סיווג קבוצות וזיהוי שיוך שגוי ({len(leaves)} עלים)...")
    for leaf in leaves:
        leaf_name = leaf["leaf_category"]
        if leaf_name in result:
            check = _detect_anomalies_for_leaf(leaf, client, model_id)
            result[leaf_name]["group_status"]   = check["group_status"]
            result[leaf_name]["anomalies"]       = check["anomalies"]
            result[leaf_name]["majority_theme"]  = check["theme"]

    return result


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

REVIEWER_COLS = [
    "technology_name",
    "Technology_Description",
    "category_level_1",
    "category_level_2",
    "category_level_3",
    "leaf_category",
    "תת-תחום חלופי",
    "רלוונטיות לתחום נוסף",
    "⚠️ חריגה אפשרית",
    "קיימת? (כן/לא/לא בטוח)",
    "שם נכון/מקובל",
    "הערות כלליות",
    "מי עובד על זה בישראל?",
]

HEADER_LABELS = {
    "technology_name": "שם הטכנולוגיה",
    "Technology_Description": "תיאור",
    "category_level_1": "קטגוריה L1",
    "category_level_2": "קטגוריה L2",
    "category_level_3": "קטגוריה L3",
    "leaf_category": "עלה סופי",
    "תת-תחום חלופי": "תת-תחום חלופי",
    "רלוונטיות לתחום נוסף": "רלוונטיות לתחום נוסף",
    "⚠️ חריגה אפשרית": "סטטוס בדיקה",
    "קיימת? (כן/לא/לא בטוח)": "קיימת?",
    "שם נכון/מקובל": "שם נכון/מקובל",
    "הערות כלליות": "הערות כלליות",
    "מי עובד על זה בישראל?": "מי עובד על זה בישראל?",
}

EMPTY_COLS = {"קיימת? (כן/לא/לא בטוח)", "שם נכון/מקובל", "הערות כלליות", "מי עובד על זה בישראל?"}
# Item-level: yellow = odd sub-topic but SAME reviewer (minor); red = belongs to a DIFFERENT reviewer (routing issue).
ITEM_YELLOW_FILL = PatternFill("solid", start_color="FFF3DA", end_color="FFF3DA")  # light amber — same reviewer
ITEM_RED_FILL    = PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0")  # light red — different reviewer
ANOMALY_FILL     = ITEM_RED_FILL  # backward-compat alias
# Group-level banners: amber = diverse (one reviewer, loose); strong red = mixed (several reviewers).
GROUP_YELLOW_FILL = PatternFill("solid", start_color="BF8F00", end_color="BF8F00")  # strong amber banner — diverse group
GROUP_RED_FILL    = PatternFill("solid", start_color="C00000", end_color="C00000")  # strong red banner — mixed group
INCOHERENT_FILL   = GROUP_RED_FILL  # backward-compat alias
CROSSREF_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")  # light gray for referral rows whose home is another reviewer

COL_WIDTHS = {
    "technology_name": 30,
    "Technology_Description": 50,
    "category_level_1": 28,
    "category_level_2": 28,
    "category_level_3": 22,
    "leaf_category": 28,
    "תת-תחום חלופי": 28,
    "רלוונטיות לתחום נוסף": 34,
    "⚠️ חריגה אפשרית": 25,
    "קיימת? (כן/לא/לא בטוח)": 18,
    "שם נכון/מקובל": 22,
    "הערות כלליות": 28,
    "מי עובד על זה בישראל?": 28,
}

HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
SUBHEAD_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
LEAF_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
DATA_FILL_A  = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
DATA_FILL_B  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
EMPTY_FILL   = PatternFill("solid", start_color="FFF9C4", end_color="FFF9C4")

THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _c(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border


DOMAIN_NAMES = {DOMAIN_1, DOMAIN_2, DOMAIN_3, DOMAIN_4}

# The model occasionally returns a domain name with a small deformation — a
# missing prefix word, or a lookalike Latin character inside a Hebrew word
# (e.g. a Latin 'g' inside "ביולוגיה"). Left uncorrected, one such character
# would send an entire reviewer file to the wrong person, so every returned
# name is folded back onto the canonical list before it is used to split files.
# The original mapping covered the real domain names and their observed
# misspellings; it is redacted here along with the names themselves.
_DOMAIN_NORM = {
    # "<observed variant>": "<canonical domain name>",
}


def _fix_domain(name):
    if not name:
        return name
    v = str(name).strip()
    # Example of the class of repair this performed (regex kept, pattern redacted):
    # v = re.sub(r"<latin-char-inside-hebrew-word>", "<canonical spelling>", v)
    return _DOMAIN_NORM.get(v, v)


def normalize_domain_map(domain_map: dict) -> dict:
    """Normalize domain names in domain_map to canonical names."""
    result = {}
    for leaf, info in domain_map.items():
        info = dict(info)
        info["primary_domain"] = _fix_domain(info.get("primary_domain"))
        if info.get("secondary_domain"):
            info["secondary_domain"] = _fix_domain(info["secondary_domain"])
        result[leaf] = info
    return result


def export_reviewer_files(
    result_df: pd.DataFrame,
    domain_map: dict,
) -> dict:
    """
    Generate one Excel file per reviewer domain.
    Returns: {domain: bytes}
    """
    df = result_df.copy()

    def _norm(val):
        if not val:
            return val
        return _fix_domain(val)

    df["primary_domain"]     = df["leaf_category"].map(lambda x: _norm(domain_map.get(x, {}).get("primary_domain", "לא מסווג")))
    df["primary_sub_domain"] = df["leaf_category"].map(lambda x: domain_map.get(x, {}).get("primary_sub_domain", ""))
    df["secondary_domain"]   = df["leaf_category"].map(lambda x: _norm(domain_map.get(x, {}).get("secondary_domain")))
    df["secondary_sub_domain"] = df["leaf_category"].map(lambda x: domain_map.get(x, {}).get("secondary_sub_domain"))
    # Mark technologies and groups by the 4-level scheme (judged at REVIEWER level):
    #   item  yellow → odd sub-topic, SAME reviewer  → "חריגה בקבוצה"
    #   item  red    → belongs to a DIFFERENT reviewer → "לא מתאימה בקבוצה"
    #   group diverse (🟡) → one reviewer, loose grouping → amber banner
    #   group mixed   (🔴) → spans several reviewers      → red banner + manual routing
    # An item flag is per-row; a group status is a property of the whole group,
    # so it gets a single banner row above the group (see the write loop) and a
    # flag in the summary sheet — never a per-technology mark.
    # name -> level ("yellow"/"red"), per leaf
    anomaly_level_map = {}
    group_status_map = {}
    for leaf, info in domain_map.items():
        status, anoms = _flag_view(info)
        group_status_map[leaf] = status
        anomaly_level_map[leaf] = {a["name"]: a["level"] for a in anoms}
    diverse_leaves = {l for l, s in group_status_map.items() if s == "diverse"}
    mixed_leaves   = {l for l, s in group_status_map.items() if s == "mixed"}

    _ITEM_YELLOW_TXT = "⚠️ טכנולוגיה חריגה בקבוצה (אותו בודק)"
    _ITEM_RED_TXT    = "⛔ טכנולוגיה לא מתאימה בקבוצה (לבדוק ולשייך ידנית לבודק)"

    def _anomaly_label(row):
        level = anomaly_level_map.get(row["leaf_category"], {}).get(row["technology_name"])
        if level == "red":
            return _ITEM_RED_TXT
        if level == "yellow":
            return _ITEM_YELLOW_TXT
        return ""

    df["⚠️ חריגה אפשרית"] = df.apply(_anomaly_label, axis=1)
    for col in EMPTY_COLS:
        df[col] = ""
    # Placeholders for the two semantic columns (filled per-copy below)
    df["תת-תחום חלופי"] = ""
    df["רלוונטיות לתחום נוסף"] = ""

    # Build per-domain dataframes.
    # A leaf with a secondary domain DIFFERENT from its primary is a true
    # cross-domain bridge: it appears in BOTH domain files. In its home
    # (primary) file it is a normal row; in the secondary file it appears as
    # a referral, sorted under the secondary sub-domain, flagged grey.
    # A secondary domain EQUAL to the primary is only an alternative
    # sub-domain — same file, shown in the "תת-תחום חלופי" column.
    domain_dfs = {}

    def _add(domain, sub, row, *, is_cross_ref, alt_sub="", cross_rel=""):
        r = row.copy()
        r["_domain"] = domain
        r["_sub_domain"] = sub or ""
        r["_is_cross_ref"] = is_cross_ref
        r["תת-תחום חלופי"] = alt_sub or ""
        r["רלוונטיות לתחום נוסף"] = cross_rel or ""
        domain_dfs.setdefault(domain, []).append(r)

    for _, row in df.iterrows():
        primary = row["primary_domain"]
        if not isinstance(primary, str) or not primary.strip():
            continue
        p_sub = row["primary_sub_domain"] if isinstance(row["primary_sub_domain"], str) else ""
        sec   = row["secondary_domain"]
        if not isinstance(sec, str) or not sec.strip():
            sec = None
        s_sub = row["secondary_sub_domain"] if isinstance(row["secondary_sub_domain"], str) else ""

        if sec and sec == primary:
            # Same domain, different sub-domain → alternative sub-domain only.
            _add(primary, p_sub, row, is_cross_ref=False, alt_sub=s_sub)
        elif sec and sec != primary:
            # True cross-domain bridge → appears in both files.
            # Home (primary) file: normal row pointing to the other domain.
            _add(primary, p_sub, row, is_cross_ref=False,
                 cross_rel=f"{sec} / {s_sub}" if s_sub else sec)
            # Secondary file: referral, sorted under the secondary sub-domain,
            # pointing back to the home domain.
            _add(sec, s_sub, row, is_cross_ref=True,
                 cross_rel=f"↪ הבית הראשי: {primary} / {p_sub}" if p_sub else f"↪ הבית הראשי: {primary}")
        else:
            # No secondary at all.
            _add(primary, p_sub, row, is_cross_ref=False)

    output_files = {}

    for domain_he in sorted(domain_dfs.keys()):
        domain_df = pd.DataFrame(domain_dfs[domain_he])
        domain_df = domain_df.sort_values(["_sub_domain", "leaf_category", "technology_name"])

        wb = Workbook()
        # Sheet 1: summary (created first, then technologies)
        ws_summary = wb.active
        ws_summary.title = "סיכום עלים"
        ws = wb.create_sheet("טכנולוגיות")
        ncols = len(REVIEWER_COLS)

        # Row 1 — title
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        t = ws["A1"]
        t.value = f"טכנולוגיות מפציעות — {domain_he}"
        _c(t,
           font=Font(name="Arial", bold=True, color="FFFFFF", size=13),
           fill=HEADER_FILL,
           align=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[1].height = 30

        # Row 2 — instructions
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        inst = ws["A2"]
        inst.value = "הוראות: מלא את 4 העמודות הצהובות לכל טכנולוגיה — האם קיימת? | שם מקובל | הערות | מי עובד על זה בישראל?"
        _c(inst,
           font=Font(name="Arial", italic=True, color="FFFFFF", size=10),
           fill=SUBHEAD_FILL,
           align=Alignment(horizontal="right", vertical="center"))
        ws.row_dimensions[2].height = 20

        # Row 3 — column headers
        for ci, col in enumerate(REVIEWER_COLS, 1):
            cell = ws.cell(row=3, column=ci, value=HEADER_LABELS.get(col, col))
            is_empty = col in EMPTY_COLS
            _c(cell,
               font=Font(name="Arial", bold=True, color="000000" if is_empty else "FFFFFF", size=10),
               fill=EMPTY_FILL if is_empty else SUBHEAD_FILL,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True),
               border=THIN)
        ws.row_dimensions[3].height = 36

        # Data rows
        prev_leaf = None
        prev_sub = None
        row_num = 4
        leaf_toggle = True

        for _, dr in domain_df.iterrows():
            leaf = dr["leaf_category"]
            sub = dr.get("_sub_domain", "") or dr.get("primary_sub_domain", "") or ""

            # Sub-domain header row when sub-domain changes
            if sub != prev_sub:
                if prev_sub is not None:
                    # Spacer row
                    for ci in range(1, ncols + 1):
                        ws.cell(row=row_num, column=ci, value="").border = THIN
                    ws.row_dimensions[row_num].height = 8
                    row_num += 1
                # Sub-domain title row — write in first cell only (no merge, RTL friendly)
                sub_cell = ws.cell(row=row_num, column=1, value=f"📂  {sub}")
                _c(sub_cell,
                   font=Font(name="Arial", bold=True, color="FFFFFF", size=11),
                   fill=SUBHEAD_FILL,
                   align=Alignment(horizontal="right", vertical="center"))
                for ci in range(2, ncols + 1):
                    filler = ws.cell(row=row_num, column=ci, value="")
                    filler.fill = SUBHEAD_FILL
                ws.row_dimensions[row_num].height = 24
                row_num += 1
                prev_sub = sub
                leaf_toggle = True

            # Leaf group separator (thin line between leaf groups within same sub-domain)
            if leaf != prev_leaf:
                if prev_leaf is not None and sub == prev_sub:
                    for ci in range(1, ncols + 1):
                        sep = ws.cell(row=row_num, column=ci, value="")
                        sep.fill = LEAF_FILL
                        sep.border = THIN
                    ws.row_dimensions[row_num].height = 4
                    row_num += 1
                # Only a MIXED group gets a banner row: its technologies are not
                # individually flagged, so the banner is its only signal. A
                # DIVERSE group is correctly routed (same reviewer) — no banner;
                # the per-item amber column already carries the detail, and a
                # banner on every loose group just made the sheet look alarming.
                _banner = None
                if leaf in mixed_leaves:
                    _banner = ("⛔ קבוצה מעורבת — לבדוק ולשייך ידנית לבודקים", GROUP_RED_FILL)
                if _banner is not None:
                    _btxt, _bfill = _banner
                    bcell = ws.cell(row=row_num, column=1, value=_btxt)
                    _c(bcell,
                       font=Font(name="Arial", bold=True, color="FFFFFF", size=10),
                       fill=_bfill,
                       align=Alignment(horizontal="right", vertical="center"))
                    for ci in range(2, ncols + 1):
                        bfill = ws.cell(row=row_num, column=ci, value="")
                        bfill.fill = _bfill
                        bfill.border = THIN
                    ws.row_dimensions[row_num].height = 22
                    row_num += 1
                leaf_toggle = not leaf_toggle

            row_fill = DATA_FILL_A if leaf_toggle else DATA_FILL_B
            is_cross_ref = bool(dr.get("_is_cross_ref", False))
            base_fill = CROSSREF_FILL if is_cross_ref else row_fill

            for ci, col in enumerate(REVIEWER_COLS, 1):
                val = dr.get(col, "")
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=row_num, column=ci, value=str(val) if val != "" else "")
                is_empty = col in EMPTY_COLS
                is_anomaly_col = col == "⚠️ חריגה אפשרית"
                is_relevance_col = col == "רלוונטיות לתחום נוסף"
                wrap = col in {"Technology_Description", "הערות כלליות", "מי עובד על זה בישראל?"}
                if is_anomaly_col and str(val).startswith("⛔"):
                    cell_fill = ITEM_RED_FILL
                elif is_anomaly_col and str(val).startswith("⚠️"):
                    cell_fill = ITEM_YELLOW_FILL
                elif is_empty:
                    cell_fill = EMPTY_FILL
                else:
                    cell_fill = base_fill
                _c(cell,
                   font=Font(name="Arial", bold=(col == "technology_name"
                                                  or (is_anomaly_col and (str(val).startswith("⚠️") or str(val).startswith("⛔")))
                                                  or (is_relevance_col and is_cross_ref)), size=10),
                   fill=cell_fill,
                   align=Alignment(vertical="top", wrap_text=wrap),
                   border=THIN)

            ws.row_dimensions[row_num].height = 55
            prev_leaf = leaf
            row_num += 1

        # Column widths + freeze
        for ci, col in enumerate(REVIEWER_COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 20)
        ws.freeze_panes = "A4"

        # Summary sheet (already created as sheet 1)
        ws2 = ws_summary
        for ci, hdr in enumerate(["עלה סופי", "תת-תחום", "מספר טכנולוגיות", "שיוך נוסף / חלופי", "סיבה (AI)"], 1):
            cell = ws2.cell(row=1, column=ci, value=hdr)
            _c(cell, font=Font(name="Arial", bold=True, color="FFFFFF", size=10),
               fill=SUBHEAD_FILL, border=THIN)
        ws2.column_dimensions["A"].width = 40
        ws2.column_dimensions["B"].width = 35
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 40
        ws2.column_dimensions["E"].width = 60

        for i, (leaf_name, count) in enumerate(
            domain_df["leaf_category"].value_counts().items(), 2
        ):
            info = domain_map.get(leaf_name, {})
            p_dom = info.get("primary_domain", "") or ""
            p_sub = info.get("primary_sub_domain", "") or ""
            s_dom = info.get("secondary_domain") or ""
            s_sub = info.get("secondary_sub_domain") or ""

            # Column B (this file's sub-domain) and column D (the extra
            # assignment) are written from THIS file's point of view.
            if s_dom and s_dom != p_dom and domain_he == s_dom:
                # This leaf is a referral into the current (secondary) file.
                sub = s_sub
                note = f"↪ הבית הראשי: {p_dom} / {p_sub}" if p_sub else f"↪ הבית הראשי: {p_dom}"
            else:
                # Current file is the home (primary).
                sub = p_sub
                if s_dom and s_dom == p_dom:
                    note = f"תת-תחום חלופי: {s_sub}" if s_sub else "תת-תחום חלופי"
                elif s_dom:
                    note = f"רלוונטי גם ל: {s_dom} / {s_sub}" if s_sub else f"רלוונטי גם ל: {s_dom}"
                else:
                    note = ""

            status = group_status_map.get(leaf_name, "tight")
            # Only a MIXED group is flagged in the summary. A diverse group is
            # correctly routed (same reviewer) and is left unmarked — colouring
            # every loose group made the summary look busy and alarming.
            if status == "mixed":
                leaf_label = f"⛔ {leaf_name}"
                note = ("⛔ קבוצה מעורבת — לבדוק ולשייך ידנית לבודקים"
                        + (f" · {note}" if note else ""))
            else:
                leaf_label = leaf_name

            c_leaf = ws2.cell(row=i, column=1, value=leaf_label)
            c_leaf.border = THIN
            c_note = ws2.cell(row=i, column=4, value=note)
            c_note.border = THIN
            if status == "mixed":
                c_leaf.fill = ITEM_RED_FILL
                c_leaf.font = Font(name="Arial", bold=True)
                c_note.fill = ITEM_RED_FILL
            ws2.cell(row=i, column=2, value=sub).border = THIN
            ws2.cell(row=i, column=3, value=count).border = THIN
            ws2.cell(row=i, column=5, value=info.get("reasoning", "")).border = THIN

        buf = io.BytesIO()
        wb.save(buf)
        output_files[domain_he] = buf.getvalue()

    return output_files
